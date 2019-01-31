import pandas as pd
import wget
from utils.go import check_enrichment
from utils.ensembl2entrez import ensembl2entrez_convertor
import time
import requests
import scipy.special
import matplotlib.pyplot as plt
from matplotlib import style
style.use("ggplot")
import scipy
import logging
sh = logging.StreamHandler()
logger = logging.getLogger("log")
logger.addHandler(sh)
from infra import *
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment
import os
from utils.hg_test import calc_HG_test
from goatools.obo_parser import GODag
from goatools.go_enrichment import GOEnrichmentStudy
from goatools.associations import read_ncbi_gene2go
from utils.clustering import plot_heatmap, find_clusters
from utils.ensembl2gene_symbol import e2g_convertor
from scipy.stats import rankdata
from utils.kmeans import kmeanssample
import seaborn as sns; sns.set(color_codes=True)
from scipy.spatial import distance
from scipy.cluster import hierarchy
from collections import defaultdict

# from goatools.associations import read_gaf
############################ () cluster and enrichment #############################



def get_cluster_classes(den, label='ivl'):
    index = 0
    color_list = []
    for i, cur in enumerate(den['color_list']):
        if i!=0 and den['color_list'][i] != den['color_list'][i-1]:
            index+=1
        color_list.append("c_{}_{}".format(den['color_list'][i], index))

    cluster_idxs = defaultdict(list)
    for c, pi in zip(color_list, den['icoord']):
        for leg in pi[1:3]:
            i = (leg - 5.0) / 10.0
            if abs(i - int(i)) < 1e-5:
                cluster_idxs[c].append(int(i))

    cluster_classes = {}
    for c, l in cluster_idxs.items():
        i_l = [den[label][i] for i in l]
        cluster_classes[c] = i_l

    return cluster_classes


# () main
def find_clusters_and_gene_enrichment(tested_gene_list_file_name, total_gene_list_file_name, gene_expression_file_name, phenotype_file_name, gene_filter_file_name=None, tested_gene_list_path=None, total_gene_list_path=None, gene_expression_path=None, phenotype_path=None, gene_filter_file_path=None, var_th_index=None, start_k=2, end_k=6, calc_go=True, enrichment_list_file_names = None, meta_groups=None, filter_expression=None, cluster_algorithm = None):
    # fetch gene expression by gene_id, divided by tumor type
    gene_sets = []
    expression_sets = []
    averaged_expression_sets = []
    tested_gene_expression = load_gene_expression_profile_by_genes(tested_gene_list_file_name, gene_expression_file_name, gene_filter_file_name, tested_gene_list_path, gene_expression_path, gene_filter_file_path)
    tested_gene_expression_headers_rows, tested_gene_expression_headers_columns, tested_gene_expression = separate_headers(tested_gene_expression)

    if filter_expression is not None:
        filtered_patients = [y for x in divided_patient_ids_by_label(phenotype_file_name, groups=filter_expression) for y in x]
        print "number of filtered patients from phenotypes: {}".format(len(filtered_patients))
    else:
        print "no filter applied"
        filtered_patients = tested_gene_expression_headers_columns

    tested_gene_expression, tested_gene_expression_headers_columns = filter_genes_dataset_by_patients(filtered_patients, tested_gene_expression_headers_columns, tested_gene_expression)
    if np.shape(tested_gene_expression)[1] == 1:
        print "no expressions were found after filtering by labels {}. skipping...".format(filter_expression)
        return None


    total_gene_list = load_gene_list(total_gene_list_file_name)
    tested_gene_list = load_gene_list(tested_gene_list_file_name)
    row_var = np.var(tested_gene_expression,axis=1)
    row_var_sorted = np.sort(row_var)[::-1]

    labels_assignment_patients = None
    if meta_groups is not None:
        print "clustering patients by groups"
        labels_assignment_patients = labels_assignments(meta_groups, phenotype_file_name,
                                               tested_gene_expression_headers_columns)

    enrichment_lists = []
    if enrichment_list_file_names is not None:
        for cur in enrichment_list_file_names:
            enrichment_lists.append(load_gene_list(cur))

    if var_th_index is None:
        var_th_index = len(row_var_sorted) - 1
    row_var_th = row_var_sorted[var_th_index]
    row_var_masked_indices = np.where(row_var_th > row_var)[0]
    gene_expression_top_var = np.delete(tested_gene_expression, row_var_masked_indices, axis=0)
    gene_expression_top_var_header_rows = np.delete(tested_gene_expression_headers_rows, row_var_masked_indices, axis=0)
    gene_expression_top_var_header_columns = tested_gene_expression_headers_columns

    clfs_results = {}
    output_rows = []
    if calc_go:
        if not os.path.exists(os.path.join(constants.GO_DIR, constants.GO_FILE_NAME)):
            wget.download(constants.GO_OBO_URL, os.path.join(constants.GO_DIR, constants.GO_FILE_NAME))
        # if not os.path.exists(os.path.join(constants.TCGA_DATA_DIR, 'goa_human.gaf')):
        #     wget.download(go_obo_url, os.path.join(constants.TCGA_DATA_DIR, 'goa_human.gaf'))
        obo_dag = GODag(os.path.join(constants.GO_DIR, constants.GO_FILE_NAME))

        assoc = read_ncbi_gene2go(os.path.join(constants.GO_DIR, constants.GO_ASSOCIATION_FILE_NAME), no_top=True)
        g = GOEnrichmentStudy([int(cur) for cur in ensembl2entrez_convertor(total_gene_list)],
                              assoc, obo_dag, methods=["bonferroni", "fdr_bh"])
        g_res = g.run_study([int(cur) for cur in ensembl2entrez_convertor(gene_expression_top_var_header_rows)])
        GO_results = [(cur.NS, cur.GO, cur.goterm.name, cur.p_uncorrected, cur.p_fdr_bh) for cur in g_res if
                      cur.p_fdr_bh <= 0.05]
        print GO_results

    if cluster_algorithm == "kmeans":

        for n_clusters in range(start_k,end_k+1):
            clfs_results[n_clusters] = []
            centres, km_clf, dist = kmeanssample(X=gene_expression_top_var, k=n_clusters, metric="euclidean")
            for i in range(n_clusters):

                ranks = []
                for j in range(n_clusters):
                    ranks.append(np.average(np.delete(gene_expression_top_var, np.where(km_clf != j)[0], axis=0)))
                ranks = rankdata(ranks)
                cluster_labels = np.array(km_clf)
                for j in range(n_clusters):
                    cluster_labels[np.where(km_clf == ranks[j] - 1)] = j
                labels_assignment = [cluster_labels + 1]

                cluster_indices = np.where(km_clf!=i)[0]
                gene_expression_cluster = np.delete(gene_expression_top_var_header_rows, cluster_indices, axis=0)
                gene_headers_row_cluster = np.delete(gene_expression_top_var_header_rows, cluster_indices, axis=0)
                clfs_results[n_clusters].append((gene_headers_row_cluster, gene_headers_row_cluster))
                desc = "k={} clustering cluster {} has {} genes".format(n_clusters, i, len(gene_expression_cluster))
                gene_list = ",".join(gene_headers_row_cluster)
                url = check_enrichment(gene_list)

                go_terms = []
                uncorrectd_pvals = []
                FDRs = []
                go_names = []
                go_ns = []
                if calc_go:
                    g_res = g.run_study([int(cur) for cur in ensembl2entrez_convertor(gene_headers_row_cluster)])
                    GO_results = [(cur.NS, cur.GO, cur.goterm.name, cur.p_uncorrected, cur.p_fdr_bh ) for cur in g_res if cur.p_fdr_bh <= 0.05]
                    if len(GO_results)>0:
                        go_ns, go_terms, go_names, uncorrectd_pvals, FDRs = zip(*GO_results)

                if len(enrichment_lists) != 0:
                    for j, cur in enumerate(enrichment_lists):
                        go_terms.append(enrichment_list_file_names[j].split(".")[0])
                        uncorrectd_pvals.append(calc_HG_test([x.split(".")[0] for x in tested_gene_list], [x.split(".")[0] for x in cur], [x.split(".")[0] for x in gene_headers_row_cluster]))
                        FDRs.append(".")
                        go_names.append(".")
                        go_ns.append(".")

                output_rows.append((desc, "\r\n".join([x.split(".")[0] for x in gene_headers_row_cluster]), url, "\r\n".join(go_ns),
                                    "\r\n".join(go_terms), "\r\n".join(go_names) , "\r\n".join(map(str, uncorrectd_pvals)), "\r\n".join(map(str, FDRs))))

        gene_sorted_heatmap = np.rot90(np.flip(gene_expression_top_var[cluster_labels.argsort(), :], 1), k=-1, axes=(1, 0))
        find_clusters(end_k, gene_sorted_heatmap, gene_expression_top_var_header_columns,
                      start_k, e2g_convertor(gene_expression_top_var_header_rows),
                      tested_gene_list_file_name ,labels_assignment=labels_assignment_patients)

        plot_heatmap(gene_expression_top_var, gene_expression_top_var_header_columns, labels_assignment,
                     gene_expression_top_var_header_rows, tested_gene_list_file_name, n_clusters=None,
                     label_index=None, phenotype_heatmap=None)

    gene_sorted_heatmap = np.rot90(np.flip(gene_expression_top_var, 1), k=-1, axes=(1, 0))
    if cluster_algorithm == "hierarchical":
        df = pd.DataFrame(data=gene_sorted_heatmap, index=gene_expression_top_var_header_columns, columns=gene_expression_top_var_header_rows)

        # correlations = df.corr()
        # correlations_array = np.asarray(df.corr())
        #
        # row_linkage = hierarchy.linkage(
        #     distance.pdist(correlations_array), method='average')
        #
        # col_linkage = hierarchy.linkage(
        #     distance.pdist(correlations_array.T), method='average')

        # enrichment_gene_list = load_gene_list("uvm_mito_part.txt")
        dct = dict(zip(np.unique(labels_assignment_patients[0]), "rbg"))
        row_colors = map(dct.get, labels_assignment_patients[0])
        dct = {1:'b', 2: 'r'}
        gene_expression_top_var_header_rows_trimmed =[x.split(".")[0] for x in gene_expression_top_var_header_rows]
        # col_colors = map(dct.get, [2 if x in enrichment_gene_list else 1 for x in gene_expression_top_var_header_rows_trimmed])
        g = sns.clustermap(df, row_colors=row_colors, metric="euclidean", robust=True, method="single")
        # den_patients = scipy.cluster.hierarchy.dendrogram(g.dendrogram_row.linkage,
        #                                          labels=df.index,
        #                                          color_threshold=0.60)
        den_genes = scipy.cluster.hierarchy.dendrogram(g.dendrogram_col.linkage,
                                                 labels=df.columns,
                                                 color_threshold=0.7)
        clusters = get_cluster_classes(den_genes)


        g.savefig(os.path.join(constants.BASE_PROFILE, "output", "hierarchical_cluster_{}.png".format(time.time())))

    for cur_labels_assignment_patient in labels_assignment_patients:
        plot_heatmap(gene_sorted_heatmap, gene_expression_top_var_header_rows, [cur_labels_assignment_patient],
                     gene_expression_top_var_header_columns, tested_gene_list_file_name, n_clusters=None,
                     label_index=None, phenotype_heatmap=None)


    print_to_excel(output_rows=output_rows, gene_list_file_name=tested_gene_list_file_name.split(".")[0], gene_expression_file_name=gene_expression_file_name.split(".")[0], var_th_index=var_th_index)


def print_to_excel(output_rows,gene_list_file_name,gene_expression_file_name,var_th_index):
    wb = Workbook()#ffff00
    ws = wb.active
    yellowFill = PatternFill(start_color='00FFFF00',
                          end_color='00FFFF00',
                          fill_type='solid')
    bd_regular = Side(style='thin', color="000000")
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)

    bd_bold = Side(style='thick', color="000000")
    border_bold = Border(left=bd_bold, top=bd_bold, right=bd_bold, bottom=bd_bold)

    blueDarkFill = PatternFill(start_color='006699FF',
                             end_color='006699FF',
                             fill_type='solid')
    blueMediumFill = PatternFill(start_color='0099CCFF',
                               end_color='0099CCFF',
                               fill_type='solid')
    blueLightFill = PatternFill(start_color='00E6F3FF',
                                 end_color='00E6F3FF',
                                 fill_type='solid')
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)


    headers = ["Description", "Genes", "URL", "GO_NS", "GO_ID", "GO_name", "nominal_pval", "FDR"]
    for i, header in enumerate(headers):
        ws['{}1'.format(chr(65+i))].border = border_regular
        ws['{}1'.format(chr(65+i))].fill = yellowFill
        ws['{}1'.format(chr(65+i))] = header
        ws.column_dimensions['{}'.format(chr(65 + i))].width = 30

    for k, cur in enumerate(headers):
        for i, cur in enumerate(output_rows):
            ws['{}{}'.format(chr(65+k), i+2)].border = border_regular
            ws['{}{}'.format(chr(65+k), i+2)].fill = blueLightFill
            ws['{}{}'.format(chr(65+k), i+2)] = cur[k]
            ws['{}{}'.format(chr(65+k), i+2)].alignment = Alignment(wrap_text=True)

    ws.column_dimensions["{}".format(chr(66+k))].width = 30
    ws["{}1".format(chr(66+k))].border = border_bold
    ws["{}1".format(chr(66+k))].fill = blueDarkFill
    ws["{}1".format(chr(66+k))] = "top var {}".format(var_th_index)
    wb.save(os.path.join(constants.OUTPUT_DIR,"CULSTERS_ENRICHMENT-{}-{}-topvar-{}-{}.xlsx".format(gene_list_file_name, gene_expression_file_name, var_th_index, time.time())))