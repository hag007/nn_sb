import re
import gzip
import shutil
import wget
from download_resources import download
from utils.ensembl2entrez import ensembl2entrez_convertor
from matplotlib import style
style.use("ggplot")
import logging
sh = logging.StreamHandler()
logger = logging.getLogger("log")
logger.addHandler(sh)
from infra import *
import os
from goatools.obo_parser import GODag
from goatools.go_enrichment import GOEnrichmentStudy
from goatools.associations import read_ncbi_gene2go
from utils.ensembl2gene_symbol import e2g_convertor
import openpyxl
import pandas as pd
import subprocess
import time
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment
from utils.stopwatch import Stopwatch

HG_GO_ROOT = "GO root"
HG_GO_ID = "GO id"
HG_GO_NAME = "GO name"
HG_PVAL = "pval"
HG_QVAL = "qval"
HG_VALUE = "value"
HG_TABLE_HEADERS = [HG_GO_ROOT, HG_GO_ID, HG_GO_NAME, HG_VALUE, HG_PVAL, HG_QVAL]

def check_enrichment(gene_list):
    ensembl_for_url = re.sub("\.\d{1,2},", ",", gene_list)
    url = "http://david.abcc.ncifcrf.gov/api.jsp?type=ENSEMBL_GENE_ID&ids={}&tool=chartReport&annot=GOTERM_BP_DIRECT,GOTERM_CC_DIRECT,GOTERM_MF_DIRECT,KEGG_PATHWAY".format(ensembl_for_url)
    return url

def check_group_enrichment(tested_gene_file_name, total_gene_file_name, algo = "", module = "", method="goatools"):
    if method=="goatools":
        return check_group_enrichment_goatools(tested_gene_file_name, total_gene_file_name)
    elif method=="tango":
        return check_group_enrichment_tango(tested_gene_file_name, total_gene_file_name, algo, module)


def check_group_enrichment_tango(tested_gene_file_name, total_gene_file_name, algo="", module=""):
    if len(tested_gene_file_name) == 0 or len(total_gene_file_name) == 0: return []

    if type(total_gene_file_name) == str:
        total_gene_list = [x.split("\t")[0] for x in load_gene_list(total_gene_file_name)]
    else:
        total_gene_list = total_gene_file_name

    if type(tested_gene_file_name) == str:
        tested_gene_list = [x.split("\t")[0] for x in load_gene_list(tested_gene_file_name)]
    else:
        tested_gene_list = tested_gene_file_name

    df_tested = pd.DataFrame(index=ensembl2entrez_convertor(tested_gene_list))
    df_tested["set"]=0
    df_tested_file_name = os.path.join(constants.OUTPUT_DIR, "_".join(["tested", algo, module]))
    df_bg_file_name = os.path.join(constants.OUTPUT_DIR, "_".join(["bg", algo, module]))
    df_tested.to_csv(df_tested_file_name, header=False, sep="\t")
    pd.DataFrame(index=ensembl2entrez_convertor(total_gene_list)).to_csv(df_bg_file_name, header=False, sep="\t")
    output_file_name = os.path.join(constants.OUTPUT_DIR, "output_{}_{}".format(algo,module))


    conf = file(os.path.join(constants.ALGO_BASE_DIR,"tango","parameter_file.format")).read().format(SET=df_tested_file_name, BACKGROUND=df_bg_file_name, OUTPUT_FILE_NAME=output_file_name)
    conf_file_name = os.path.join(constants.OUTPUT_DIR, "parameter_file_{}_{}_{}".format(algo,module,time.time()))
    file(conf_file_name,'w+').write(conf)

    print subprocess.Popen("wine win/annot_sets.exe {}".format(conf_file_name), shell=True,
                           stdout=subprocess.PIPE, cwd=os.path.join(constants.ALGO_BASE_DIR, "tango")).stdout.read()

    df_results = pd.DataFrame()
    if os.path.isfile(output_file_name) and os.path.getsize(output_file_name) > 1:
        df_results = pd.read_csv(output_file_name, sep="\t", index_col=False, header=None)

    hg_report = []
    go_terms = []
    uncorrectd_pvals = []
    FDRs = []
    go_names = []
    go_ns = []
    if len(df_results.index) > 0:
        # go_ns, go_terms, go_names, go_hg_value, uncorrectd_pvals, FDRs = zip(*[("NA", cur[1]["Category"].split(" - ")[1], cur[1]["Category"].split(" - ")[0], cur[1]["Gene IDs"].count(',')+1, cur[1]["Raw Pvalue"], cur[1]["p-value"]) for cur in df_results.iterrows()])
        # hg_report = [{HG_GO_ROOT: "NA", HG_GO_ID: cur[1]["Category"].split(" - ")[1], HG_GO_NAME: cur[1]["Category"].split(" - ")[0], HG_VALUE: cur[1]["Gene IDs"].count(',')+1, HG_PVAL: cur[1]["Raw Pvalue"],
        #               HG_QVAL: cur[1]["p-value"]} for cur in df_results.iterrows()]
        go_ns, go_terms, go_names, go_hg_value, uncorrectd_pvals, FDRs = zip(*[("NA", cur[1][6], cur[1][1], cur[1][4], 10**float(cur[1][2]), 10**float(cur[1][3])) for cur in df_results.iterrows()])
        hg_report = [{HG_GO_ROOT: "NA", HG_GO_ID: cur[1][6], HG_GO_NAME: cur[1][1], HG_VALUE: cur[1][5], HG_PVAL: 10**float(cur[1][2]),
                      HG_QVAL: 10**float(cur[1][3])} for cur in df_results.iterrows()]
        hg_report.sort(key=lambda x: x[HG_QVAL])
        hg_report=filter(lambda x: x[HG_QVAL] <=0.05, hg_report)

    output_rows = [("\r\n".join(e2g_convertor(tested_gene_list)),  "\r\n".join(go_ns),
                        "\r\n".join(go_terms), "\r\n".join(go_names), "\r\n".join(map(str, uncorrectd_pvals)),
                        "\r\n".join(map(str, FDRs)))]
    print_to_excel(output_rows, str(tested_gene_file_name)[:10], str(total_gene_file_name)[:10])
    return hg_report



def check_group_enrichment_goatools(tested_gene_file_name, total_gene_file_name, th=1):
    if len(tested_gene_file_name) == 0 or len(total_gene_file_name) == 0: return []

    if type(total_gene_file_name) == str:
        total_gene_list = load_gene_list(total_gene_file_name)
    else:
        total_gene_list = total_gene_file_name

    if type(tested_gene_file_name) == str:
        tested_gene_list = load_gene_list(tested_gene_file_name)
    else:
        tested_gene_list = tested_gene_file_name

    if not os.path.exists(os.path.join(constants.GO_DIR, constants.GO_FILE_NAME)):
        download(constants.GO_OBO_URL, constants.GO_DIR)

    obo_dag = GODag(os.path.join(constants.GO_DIR, constants.GO_FILE_NAME))

    if not os.path.exists(os.path.join(constants.GO_DIR, constants.GO_ASSOCIATION_FILE_NAME)):
        download(constants.GO_ASSOCIATION_GENE2GEO_URL, constants.GO_DIR)
        with gzip.open(os.path.join(constants.GO_DIR, os.path.basename(constants.GO_ASSOCIATION_GENE2GEO_URL)), 'rb') as f_in:
            with open(os.path.join(constants.GO_DIR, constants.GO_ASSOCIATION_FILE_NAME),'wb') as f_out:
                shutil.copyfileobj(f_in, f_out)

    assoc = read_ncbi_gene2go(os.path.join(constants.GO_DIR, constants.GO_ASSOCIATION_FILE_NAME), no_top=True)

    sw=Stopwatch()
    sw.start()
    g = GOEnrichmentStudy([int(cur) for cur in ensembl2entrez_convertor(total_gene_list)],
                          assoc, obo_dag, methods=[], log=None) # "bonferroni", "fdr_bh"
    g_res = g.run_study([int(cur) for cur in ensembl2entrez_convertor(tested_gene_list)])
    print sw.stop("done GO analysis in ")
    # GO_results = [(cur.NS, cur.GO, cur.goterm.name, cur.pop_count, cur.p_uncorrected, cur.p_fdr_bh) for cur in g_res if
    #               cur.p_fdr_bh <= 0.05]
    GO_results = [(cur.NS, cur.GO, cur.goterm.name, cur.pop_count, cur.p_uncorrected) for cur in g_res if
                  cur.p_uncorrected <= th]

    hg_report = [{HG_GO_ROOT : cur[0], HG_GO_ID : cur[1], HG_GO_NAME : cur[2], HG_VALUE : cur[3], HG_PVAL : cur[4] , HG_QVAL : 1} for cur in GO_results] # , HG_QVAL : cur[5]
    # hg_report.sort(key=lambda x: x[HG_QVAL])
    hg_report.sort(key=lambda x: x[HG_PVAL])

    if len(GO_results) > 0:
        go_ns, go_terms, go_names, go_hg_value, uncorrectd_pvals = zip(*GO_results) # , FDRs
    else:
        go_terms = []
        uncorrectd_pvals = []
        FDRs = []
        go_names = []
        go_ns = []
    # output_rows = [("\r\n".join(e2g_convertor(tested_gene_list)),  "\r\n".join(go_ns),
    #                     "\r\n".join(go_terms), "\r\n".join(go_names), "\r\n".join(map(str, uncorrectd_pvals)),
    #                     "\r\n".join(map(str, FDRs)))]
    # print_to_excel(output_rows, str(tested_gene_file_name)[:10], str(total_gene_file_name)[:10])
    return hg_report




def print_to_excel(output_rows, gene_list_file_name, total_gene_file_name):
    wb = Workbook()  # ffff00
    ws = wb.active
    yellowFill = PatternFill(start_color='00FFFF00',
                             end_color='00FFFF00',
                             fill_type='solid')
    bd_regular = Side(style='thin', color="000000")
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)

    bd_bold = Side(style='thick', color="000000")
    border_bold = Border(left=bd_bold, top=bd_bold, right=bd_bold, bottom=bd_bold)

    blueDarkFill = PatternFill(start_color='006699FF',
                               end_color='006699FF',
                               fill_type='solid')
    blueMediumFill = PatternFill(start_color='0099CCFF',
                                 end_color='0099CCFF',
                                 fill_type='solid')
    blueLightFill = PatternFill(start_color='00E6F3FF',
                                end_color='00E6F3FF',
                                fill_type='solid')
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)

    headers = ["Genes", "GO_NS", "GO_ID", "GO_name", "nominal_pval", "FDR"]
    for i, header in enumerate(headers):
        ws['{}1'.format(chr(65 + i))].border = border_regular
        ws['{}1'.format(chr(65 + i))].fill = yellowFill
        ws['{}1'.format(chr(65 + i))] = header
        ws.column_dimensions['{}'.format(chr(65 + i))].width = 30

    for k, cur in enumerate(headers):
        for i, cur in enumerate(output_rows):
            ws['{}{}'.format(chr(65 + k), i + 2)].border = border_regular
            ws['{}{}'.format(chr(65 + k), i + 2)].fill = blueLightFill
            ws['{}{}'.format(chr(65 + k), i + 2)] = cur[k]
            ws['{}{}'.format(chr(65 + k), i + 2)].alignment = Alignment(wrap_text=True)

    ws.column_dimensions["{}".format(chr(66 + k))].width = 30
    ws["{}1".format(chr(66 + k))].border = border_bold
    ws["{}1".format(chr(66 + k))].fill = blueDarkFill
    wb.save(os.path.join(constants.OUTPUT_DIR,
                         "GENE_SET_ENRICHMENT-{}-{}-{}.xlsx".format(gene_list_file_name.split(".")[0], total_gene_file_name.split(".")[0],
                                                                               time.time())))
