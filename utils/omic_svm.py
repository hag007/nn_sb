import sys
import mord
import scipy.special
from matplotlib import style
style.use("ggplot")
from sklearn import svm
from sklearn.model_selection import GridSearchCV, cross_val_score
import scipy
from scipy.stats import hypergeom
import logging
sh = logging.StreamHandler()
logger = logging.getLogger("log")
logger.addHandler(sh)
from constants import *
from infra import *
from sklearn.metrics import precision_recall_curve
from sklearn.metrics import average_precision_score
from sklearn.metrics import roc_auc_score
from sklearn.metrics import roc_curve
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment
from utils.groups_generator import *
import math
import time

DISTANCE = "distance"
LOGISTIC_REGRESSION = "logistic_regression"


############################################(3) SVM to see whether we can predict tumor type by tested genes' profile expression########################################

def load_svm_data(tested_gene_file_name, expression_profile_file_name, phenotype_file_name, label=None, label_values=None, gene_filter_file_name=None, groups=None):
    data = []
    labels = []
    labeled_groups = load_expression_profile_by_labelling(tested_gene_file_name, expression_profile_file_name, phenotype_file_name, label, label_values, gene_filter_file_name, groups=groups)
    print "groups sizes: {}".format([len(cur_group) for cur_group in labeled_groups])
    labeled_groups = sorted(labeled_groups, key = lambda x: len(x), reverse=True)
    # min_size = min([len(cur) for cur in labeled_groups])
    # labeled_groups = [cur[:min_size] for cur in labeled_groups]
    for i, cur_group in enumerate(labeled_groups):
        for cur in cur_group[1:]:
            data.append(cur[1:].astype(np.float))
            labels.append(i)


    return data, labels, labeled_groups, labeled_groups[0][0][1:]

def randonize_patients(data, labels):

    while True:
        total  = zip(data, labels)
        random.shuffle(total)
        data, label = zip(*total)
        if sum(label[(3 * len(label)) / 4:]) % len(label[(3 * len(label)) / 4:]) != 0 and \
            sum(label[:(3 * len(label)) / 4]) % len(label[:(3 * len(label)) / 4]) != 0:
            break
    return (data, label)
def randonize_patients_old(data, labels):
    total = np.c_[data, labels]
    random.shuffle(total)
    # results = zip(*total)
    return total[:,:-1], total[:,-1]


def divide_train_and_test_groups(data, labels):
    data_train = data[:-1] # data[:(3 * len(data)) / 4]
    data_test = data[-1:] # data[(3 * len(data)) / 4:]
    labels_train = labels[:-1] # labels[:(3 * len(data)) / 4]
    labels_test = labels[-1:] # labels[(3 * len(data)) / 4:]
    return data_train, data_test, labels_train, labels_test

def svm_rbf_default(tuned_parameters):
    return GridSearchCV(svm.SVC(probability=True), param_grid=tuned_parameters, return_train_score=True)

def svm_linear_default(tuned_parameters):
    return GridSearchCV(svm.SVC(probability=True, kernel='linear'), param_grid=tuned_parameters, return_train_score=True)

def svm_multiclass(tuned_parameters):
    return mord.LogisticAT(alpha=1.)

def apply_svm(clf_method, data_train, labels_train, data_test, labels_test, rank_method, labels_permutation=LABELS_NORMAL):

    data_train = [[cur2 for cur2 in cur1] for cur1 in data_train]
    labels_train = [cur for i, cur in enumerate(labels_train)]

    data_test = [[cur2 for cur2 in cur1] for cur1 in data_test]
    labels_test = [cur for i, cur in enumerate(labels_test)]

    if labels_permutation == constants.LABELS_RANDOM:
        labels_train = [math.floor(random.random() / 0.5) for i, cur in enumerate(labels_train)]
        labels_test = [math.floor(random.random() / 0.5) for i, cur in enumerate(labels_test)]
    if labels_permutation == constants.LABELS_SHUFFLE:
        random.shuffle(labels_train)
        random.shuffle(labels_test)
    if labels_permutation == constants.LABELS_ALTERNATED:
        labels_train = [i % 2 for i, cur in enumerate(labels_train)]
        labels_test = [i % 2 for i, cur in enumerate(labels_test)]
    if labels_permutation == constants.LABELS_INVERTED:
        labels_train = [abs(cur - 1) for i, cur in enumerate(labels_train)]
        labels_test = [abs(cur - 1) for i, cur in enumerate(labels_test)]

    data_train=np.array(data_train)
    labels_train=np.array(labels_train)
    data_test = np.array(data_test)
    labels_test = np.array(labels_test)
    clf_method.fit(data_train, labels_train)
    predicted_results = clf_method.predict(data_test)
    if rank_method == DISTANCE:
        probabilities = clf_method.decision_function(data_test)
    else:
        probabilities = [cur[predicted_results[i]] for i, cur in enumerate(clf_method.predict_proba(data_test))]
        # probabilities = probabilities[:,1]
    #zipped = zip(probabilities, data_train, labels_train)
    #zipped_sorted = sorted(zipped, key=lambda x: x[0])
    #patients_rank_sorted = [x[0] for x in zipped_sorted]
    #patients_expression_sorted = [x[1] for x in zipped_sorted]
    #patients_labels_sorted = [x[2] for x in zipped_sorted]
    precision, recall, _ = precision_recall_curve(labels_test, probabilities)
    average_precision = average_precision_score(labels_test, probabilities)
    fpr, tpr, _ = roc_curve(labels_test, probabilities)
    auc = roc_auc_score(labels_test, probabilities) # int(labels_test[0] == predicted_results[0])


    # prediction = clf_method.predict(data_test)
    # predicted_positives = sum(prediction)
    # predicted_negatives = len(prediction) - predicted_positives
    # labeled_positives = sum(labels_test)
    # labeled_negatives = len(labels_test) - labeled_positives
    #
    # tp = []
    # fp = []
    # fn = []
    # tn = []
    # precision_1 = []
    # recall_1 = []
    # for ind, cur in enumerate(prediction):
    #     labeled_positives = sum(labels_test[:ind])
    #     labeled_negatives = len(labels_test[:ind]) - labeled_positives
    #     tp.append(sum([1 for i, cur in enumerate(prediction[:ind]) if cur == 1 and labels_test[i] == 1]))
    #     fp.append(sum([1 for i, cur in enumerate(prediction[:ind]) if cur == 1 and labels_test[i] == 0]))
    #     fn.append(labeled_positives - tp[-1])
    #     tn.append(labeled_negatives - fp[-1])
    #     precision_1.append(float(tp[-1]) / (tp[-1] + fn[-1]))
    #     recall_1.append(float(tp[-1]) / (tp[-1] + fp[-1]))
    #
    # recall_1, precision_1 = zip(*sorted(zip(recall, precision), key=lambda x: x[0]))
    #
    # accuracy = 1 - float(sum(abs(prediction - labels_test)))/len(labels_test)

    print "PR: {}, ROC: {}".format(average_precision,auc) ## , fp: {}, fn: {}, total: {}  ## , fp[-1], fn[-1], len(prediction))

    # plt.plot(recall_1, precision_1, label="PRAUC")
    # plt.step(recall, precision, color='b', alpha=0.2,
    #          where='post')
    # plt.fill_between(recall, precision, step='post', alpha=0.3,
    #                  color='b')
    #
    # plt.xlabel('Recall')
    # plt.ylabel('Precision')
    # plt.ylim([0.0, 1.05])
    # plt.xlim([0.0, 1.0])
    # plt.title('2-class Precision-Recall curve: AP={0:0.2f}'.format(average_precision))
    #
    # plt.show()
    # plt.savefig(os.path.join(BASE_OUTPUT_DIR,"test.png"))
    #################
    # plt.plot(fpr, tpr, label="PRAUC")
    # plt.step(fpr, tpr, color='b', alpha=0.2,
    #          where='post')
    # plt.fill_between(fpr, tpr, step='post', alpha=0.3,
    #                  color='b')
    #
    # plt.xlabel('fpr')
    # plt.ylabel('tpr')
    # plt.ylim([0.0, 1.05])
    # plt.xlim([0.0, 1.0])
    # plt.title('2-class Precision-Recall curve: AP={0:0.2f}'.format(auc))
    #
    # plt.show()
    # plt.savefig(os.path.join(BASE_OUTPUT_DIR,"test.png"))


    #################
    ##
    # ta = (len(labels_test) * 1.0 - sum([abs(p - r) for p, r in zip(predicted_results, labels_test)])) / (len(labels_test) * 1.0)
    # return clf_method.best_score_, ta
    return average_precision, auc, clf_method

def print_svm_results(train_scores, train_alt_scores, test_scores, test_alt_scores, rounds):
    print "train_accuracy_avg: {}".format(sum(train_scores)/ rounds)
    print "train_alt_accuracy_avg: {}".format(sum(train_alt_scores)/ rounds)
    print "train_accuracy_diff_avg: {}".format((sum(train_scores)-sum(train_alt_scores))/rounds)
    print "test_accuracy_avg: {}".format(sum(test_scores)/ rounds)
    print "test_alt_accuracy_avg: {}".format(sum(test_alt_scores)/ rounds)
    print "test_accuracy_diff_avg: {}".format((sum(test_scores)-sum(test_alt_scores))/rounds)
    print "train p val: {}".format(scipy.stats.ttest_ind(train_scores, train_alt_scores)[1])
    print "test_p_val: {}".format(scipy.stats.ttest_ind(test_scores, test_alt_scores)[1])

    # print "{}".format(train_score_sum / rounds)
    # print "{}".format(train_alt_score_sum / rounds)
    # print "{}".format(scipy.stats.ttest_ind(train_scores, train_alt_scores)[1])
    # print "{}".format(test_score_sum / rounds)
    # print "{}".format(test_alt_score_sum / rounds)
    # print "{}".format(scipy.stats.ttest_ind(test_scores, train_alt_scores)[1])

def print_svm_results(test_scores, rounds):
    avgs = []
    print "start avg and var:"
    for i, cur_i in enumerate(test_scores):
        avg = sum(test_scores[i]) / rounds
        print "{}\t {}".format(avg, np.var(test_scores[i]))
        avgs.append(avg)
    print "done PR-AUC: avg and var"

    results_summarized = []
    for i, cur_i in enumerate(avgs):
        results_summarized.append([])

    for i, cur_i in enumerate(avgs):
        for j, cur_j in enumerate(avgs):
            if j>i:
                pval = scipy.stats.ttest_ind(test_scores[i], test_scores[j])[1]
                #print "test_p_val for {}, {}: {}".format(i,j,pval)
                results_summarized[i].append(pval)
            elif j==i:
                results_summarized[i].append(1)
            else:
                results_summarized[i].append(-1)

    for i, cur_i in enumerate(avgs):
        for j, cur_j in enumerate(avgs):
            if results_summarized[i][j] == -1:
                results_summarized[i][j] = results_summarized[j][i]

    # for i, cur_i in enumerate(avgs):
    #     print ""
    #     for j, cur_j in enumerate(avgs):
    #         print "{}\t".format(results_summarized[i][j]),

    # print "{}".format(train_score_sum / rounds)
    # print "{}".format(train_alt_score_sum / rounds)
    # print "{}".format(scipy.stats.ttest_ind(train_scores, train_alt_scores)[1])
    # print "{}".format(test_score_sum / rounds)
    # print "{}".format(test_alt_score_sum / rounds)
    # print "{}".format(scipy.stats.ttest_ind(test_scores, train_alt_scores)[1])


# (3) main
def prediction_by_gene_expression(gene_list_file_names, gene_expression_file_name, phenotype_file_name, label=None, label_values=None, rank_method=LOGISTIC_REGRESSION, gene_filter_file_name=None, rounds=2, groups=None, classification_method="svm_rbf_default", tuning_parameters={'C': [10], 'kernel': ['rbf']}, labels_permutation=constants.LABELS_NORMAL, compare_to_random=True):
    thismodule = sys.modules[__name__]
    clf = getattr(thismodule, classification_method)(tuning_parameters)
    genelist_datasets = []
    gene_list_sizes = []
    for tested_gene_file_name in gene_list_file_names:
        data, labels , _1, _2 = load_svm_data(tested_gene_file_name, gene_expression_file_name, phenotype_file_name, label, label_values,
                                              gene_filter_file_name, groups)
        genelist_datasets.append(data)
        gene_list_sizes.append(np.shape(data)[1])

    if compare_to_random:
        for cur_size in list(gene_list_sizes):
            random_set_file_name = generate_random_set(random_size = cur_size, meta_gene_set="protein_coding_long.txt")
            data, labels, _1, _2 = load_svm_data(random_set_file_name, gene_expression_file_name, phenotype_file_name,
                                                 label, label_values,
                                                 gene_filter_file_name, groups)
            genelist_datasets.append(data)
            gene_list_sizes.append(np.shape(data)[1])
            gene_list_file_names.append(random_set_file_name)

    train_scores = []
    test_pr_score = []
    test_roc_score = []
    for j in range(len(genelist_datasets)):
        train_scores.append([])
        test_pr_score.append([])
        test_roc_score.append([])
    for i in range(rounds):
        genelist_datasets = np.rot90(genelist_datasets, k=1, axes=(1, 0))
        genelist_datasets, labels = randonize_patients(genelist_datasets, labels)
        genelist_datasets = np.rot90(genelist_datasets, k=-1, axes=(1, 0))
        for j, cur_dataset in enumerate(genelist_datasets):
            data_train, data_test, labels_train, labels_test = divide_train_and_test_groups(cur_dataset, labels)

            test_pr, test_roc = apply_svm(clf, data_train, labels_train, data_test, labels_test, rank_method, labels_permutation)
            test_pr_score[j].append(test_pr)
            test_roc_score[j].append(test_roc)
    print "#######################################"
    print "PRAUC"
    print_svm_results(test_pr_score, float(rounds))
    print "ROC"
    print_svm_results(test_roc_score, float(rounds))
    print_to_excel(gene_sets_names=[cur.split('.')[0] for cur in gene_list_file_names], gene_sets_sizes=gene_list_sizes, results=(test_pr_score, test_roc_score), rank_method=rank_method)



def make_meshgrid(x, y, h=.02):
    """Create a mesh of points to plot in

    Parameters
    ----------
    x: data to base x-axis meshgrid on
    y: data to base y-axis meshgrid on
    h: stepsize for meshgrid, optional

    Returns
    -------
    xx, yy : ndarray
    """
    x_min, x_max = x.min() - 1, x.max() + 1
    y_min, y_max = y.min() - 1, y.max() + 1
    xx, yy = np.meshgrid(np.arange(x_min, x_max, h),
                         np.arange(y_min, y_max, h))
    return xx, yy


def plot_contours(ax, clf, xx, yy, **params):
    """Plot the decision boundaries for a classifier.

    Parameters
    ----------
    ax: matplotlib axes object
    clf: a classifier
    xx: meshgrid ndarray
    yy: meshgrid ndarray
    params: dictionary of params to pass to contourf, optional
    """
    Z = clf.predict(np.c_[xx.ravel(), yy.ravel()])
    Z = Z.reshape(xx.shape)
    out = ax.contourf(xx, yy, Z, **params)
    return out


def print_to_excel(gene_sets_names, gene_sets_sizes, results=None, rank_method="LOGISTIC_REGRESSION"):
    wb = Workbook()#ffff00
    ws = wb.active
    yellowFill = PatternFill(start_color='00FFFF00',
                          end_color='00FFFF00',
                          fill_type='solid')
    bd_regular = Side(style='thin', color="000000")
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)

    bd_bold = Side(style='thick', color="000000")
    border_bold = Border(left=bd_bold, top=bd_bold, right=bd_bold, bottom=bd_bold)

    ws.column_dimensions["A"].width = 20

    ws['A1'].border = border_regular
    ws['A1'].fill = yellowFill
    ws['B1'].border = border_regular
    ws['B1'].fill = yellowFill
    for i in range(1,len(gene_sets_names)+1):
        if (65+i*2) > 90:
            prefix = "A"
        else:
            prefix = ""
        ws.merge_cells('{}{}1:{}{}1'.format(prefix,chr(65+(i*2)%26),prefix,chr(65+(i*2+1)%26)))
        ws['{}{}1'.format(prefix,chr(65+(i*2)%26))].border = border_regular
        ws['{}{}1'.format(prefix,chr(65+(i*2+1)%26))].border = border_regular
        ws['{}{}1'.format(prefix,chr(65+(i*2)%26))].fill = yellowFill
        ws['{}{}1'.format(prefix,chr(65+(i*2)%26))] = gene_sets_names[i-1] + " (n={})".format(gene_sets_sizes[i-1])
        ws['{}{}1'.format(prefix,chr(65+(i*2)%26))].alignment = Alignment(horizontal='center', wrap_text=True)

    blueDarkFill = PatternFill(start_color='006699FF',
                             end_color='006699FF',
                             fill_type='solid')
    blueMediumFill = PatternFill(start_color='0099CCFF',
                               end_color='0099CCFF',
                               fill_type='solid')
    blueLightFill = PatternFill(start_color='00E6F3FF',
                                 end_color='00E6F3FF',
                                 fill_type='solid')
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)
    ws['A2'].border = border_regular
    ws['A2'].fill = blueDarkFill
    ws['B2'].border = border_regular
    ws['B2'].fill = blueMediumFill
    for i in range(1,len(gene_sets_names)+1):
        if (65+i*2) > 90:
            prefix = "A"
        else:
            prefix = ""
        ws['{}{}2'.format(prefix,chr(65+(i*2)%26))].border = border_regular
        ws['{}{}2'.format(prefix,chr(65+(i*2)%26))].fill = blueMediumFill
        ws['{}{}2'.format(prefix,chr(65+(i*2)%26))] = "avg"
        ws['{}{}2'.format(prefix,chr(65+(i*2)%26))].alignment = Alignment(horizontal='center')

        ws['{}{}2'.format(prefix,chr(65+(i*2+1)%26))].border = border_regular
        ws['{}{}2'.format(prefix,chr(65+(i*2+1)%26))].fill = blueMediumFill
        ws['{}{}2'.format(prefix,chr(65+(i*2+1)%26))] = "var"
        ws['{}{}2'.format(prefix,chr(65+(i*2+1)%26))].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:A4')
    ws['A3'].border = border_regular
    ws['A3'].fill = blueDarkFill
    ws['A4'].border = border_regular
    ws['A4'].fill = blueDarkFill
    ws['A3'] = rank_method
    ws['A3'].alignment = Alignment(horizontal='center')

    ws['B3'].border = border_regular
    ws['B3'].fill = blueMediumFill
    ws['B3'] = "PR"
    ws['B3'].alignment = Alignment(horizontal='center')
    ws['B4'].border = border_regular
    ws['B4'].fill = blueMediumFill
    ws['B4'] = "ROC"
    ws['B4'].alignment = Alignment(horizontal='center')

    for k in range(len(list(results))):
        for i in range(1,len(gene_sets_names)+1):
            if (65 + i * 2) > 90:
                prefix = "A"
            else:
                prefix = ""
            ws['{}{}{}'.format(prefix,chr(65+(i*2)%26),3+k)].border = border_regular
            ws['{}{}{}'.format(prefix,chr(65+(i*2)%26),3+k)].fill = blueLightFill
            ws['{}{}{}'.format(prefix,chr(65+(i*2)%26),3+k)] = sum(results[k][i-1])/ len(results[k][i-1])
            ws['{}{}{}'.format(prefix,chr(65+(i*2)%26),3+k)].alignment = Alignment(horizontal='center')
            ws['{}{}{}'.format(prefix,chr(65+(i*2)%26), 3 + k)].number_format = '0.000'
            ws['{}{}{}'.format(prefix,chr(65+(i*2)%26+1),3+k)].border = border_regular
            ws['{}{}{}'.format(prefix,chr(65+(i*2)%26+1),3+k)].fill = blueLightFill
            ws['{}{}{}'.format(prefix,chr(65+(i*2)%26+1),3+k)] = np.var((results[k][i-1]))
            ws['{}{}{}'.format(prefix,chr(65+(i*2)%26+1),3+k)].alignment = Alignment(horizontal='center')
            ws['{}{}{}'.format(prefix,chr(65+(i*2)%26+1), 3 + k)].number_format = '0.0000'


    ws['{}{}{}'.format(prefix,chr(65+(i*2)%26+2), 4 + k)].border = border_bold
    ws['{}{}{}'.format(prefix,chr(65+(i*2)%26+2), 4 + k)].fill = yellowFill
    ws['{}{}{}'.format(prefix,chr(65+(i*2)%26+2), 4 + k)] = "rounds = {}".format(len(results[k][i-1]))
    ws['{}{}{}'.format(prefix,chr(65+(i*2)%26+2), 4 + k)].alignment = Alignment(horizontal='center')
    ws['{}{}{}'.format(prefix,chr(65+(i*2)%26+2), 4 + k)].number_format = '0.0000'

    wb.save(os.path.join(constants.OUTPUT_DIR,"SVM_AUC-{}-{}.xlsx".format(rank_method, time.time())))
